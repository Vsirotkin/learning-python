# to get connected to worksheet in exel via python file

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# wb = load_workbook('working_with_python.xlsx')
# ws = wb.active
# print(ws['B2'].value)
#
# wb.create_sheet('from python')
# print(wb.sheetnames)

wb = Workbook()
ws = wb.active
ws.title = 'Data'

# ws.append(['Term', 'Data', 'Price', 'Manager'])
# wb.save('tim.xlsx')

for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)

wb.save('col and range.xlsx')
